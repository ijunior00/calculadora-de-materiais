from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from starlette.responses import Response
from pydantic import BaseModel
from fpdf import FPDF
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Optional, Any, Dict
import os
import time
import re
import json
import uuid
import base64
import secrets
from datetime import datetime, timezone

app = FastAPI()

# ── Authentication ────────────────────────────────────────────────
# HTTP Basic Auth gated by the APP_PASSWORD env var. When it is unset
# (e.g. local development), the app stays open. On a public host like
# Render, set APP_PASSWORD (and optionally APP_USERNAME) to lock it.
APP_USERNAME = os.environ.get("APP_USERNAME", "vestas")
APP_PASSWORD = os.environ.get("APP_PASSWORD", "")


@app.middleware("http")
async def _basic_auth(request, call_next):
    if APP_PASSWORD:
        header = request.headers.get("Authorization", "")
        authorized = False
        if header.startswith("Basic "):
            try:
                user, _, pw = base64.b64decode(header[6:]).decode("utf-8").partition(":")
                authorized = (
                    secrets.compare_digest(user, APP_USERNAME)
                    and secrets.compare_digest(pw, APP_PASSWORD)
                )
            except Exception:
                authorized = False
        if not authorized:
            return Response(
                content="Authentication required",
                status_code=401,
                headers={"WWW-Authenticate": 'Basic realm="BRMP"'},
            )
    return await call_next(request)


# ── CORS ──────────────────────────────────────────────────────────
# Off by default (frontend and backend are same-origin). Set
# ALLOWED_ORIGINS to a comma-separated list only if you ever split the
# frontend onto a different host.
_origins = [o.strip() for o in os.environ.get("ALLOWED_ORIGINS", "").split(",") if o.strip()]
if _origins:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=_origins,
        allow_methods=["GET", "POST"],
        allow_headers=["*"],
    )

# Stamped once when the server process starts. Appended as ?v=<bust> to every
# static asset URL so the browser always fetches fresh JS/CSS after a restart —
# without clearing any user data (passwords, cookies, history stay untouched).
_CACHE_BUST = str(int(time.time()))

app.mount("/static", StaticFiles(directory="static"), name="static")

def _serve_html(path: str) -> Response:
    """Read an HTML file, cache-bust its static asset URLs, and return it with
    no-cache headers so a server restart always ships fresh JS/CSS."""
    with open(path, "r", encoding="utf-8") as f:
        html = f.read()
    # Match both "static/..." and "/static/..." paths in script src and link href
    html = re.sub(r'(<script[^>]+src=")(/?static/[^"]+)(")', rf'\1\2?v={_CACHE_BUST}\3', html)
    html = re.sub(r'(<link[^>]+href=")(/?static/[^"]+)(")', rf'\1\2?v={_CACHE_BUST}\3', html)
    return Response(
        content=html,
        media_type="text/html",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"}
    )


@app.get("/", response_class=HTMLResponse)
async def get_index():
    return _serve_html("front-end/index.html")


@app.get("/m", response_class=HTMLResponse)
async def get_mobile():
    """Simplified mobile UI — same engine, fewer inputs, phone-friendly."""
    return _serve_html("front-end/mobile.html")


class BOMPayload(BaseModel):
    turbine_model: str
    blade_type: str
    damage_cat: str
    blade_zone: str
    rstart: float
    rend: float
    length: float
    width: float
    days: int
    estimated_days: Optional[int] = None
    is_external: Optional[bool] = None
    report_title: Optional[str] = ""
    doc_refs: Optional[Dict[str, Any]] = None
    include_field_rules: Optional[bool] = False
    total_brl: float = 0
    total_eur: float = 0
    blade_sn: Optional[str] = ""
    service_order: Optional[str] = ""
    cir_number: Optional[str] = ""
    damage_description: Optional[str] = ""
    chord_ref: Optional[str] = "LE"
    x1: Optional[float] = 0
    items: list
    # Optional audit payload — frontend supplies the layers, repair_steps and
    # layup summary so each PDF generation has a matching JSON record. Not
    # exposed via any HTTP route; used only by the audit writer.
    audit_inputs: Optional[Dict[str, Any]] = None
    audit_outputs: Optional[Dict[str, Any]] = None


# ── Color palette ─────────────────────────────────────────────
# ONLY the top header bar is blue. Everything else is black/grey/white.
HEADER_BLUE     = (20,  58,  95)
HEADER_TEXT     = (255, 255, 255)
BLACK           = (0,   0,   0)
DARK_TEXT       = (30,  30,  30)
MID_TEXT        = (80,  80,  80)
LIGHT_TEXT      = (130, 130, 130)
ROW_ALT         = (248, 248, 248)   # very subtle alternating grey
ROW_WHITE       = (255, 255, 255)
COL_HDR_BG      = (235, 235, 235)   # column header: light grey
CAT_HDR_BG      = (225, 225, 225)   # category header: slightly darker
LINE_COLOR      = (0,   0,   0)     # ALL table lines are black
LINE_W          = 0.15              # fine hairline


# Replacement map for characters outside Latin-1 that commonly appear in
# user-entered text or engine-generated descriptions.
_LATIN1_SUBS = str.maketrans({
    '\u2014': '-',   # em dash
    '\u2013': '-',   # en dash
    '\u2018': "'",   # left single quotation mark
    '\u2019': "'",   # right single quotation mark
    '\u201c': '"',   # left double quotation mark
    '\u201d': '"',   # right double quotation mark
    '\u00b2': '2',   # superscript 2 (m²)
    '\u00b0': 'deg', # degree sign
    '\u00b1': '+/-', # plus-minus
    '\u00e9': 'e',   # é
    '\u00e3': 'a',   # ã
    '\u00ea': 'e',   # ê
    '\u00e7': 'c',   # ç
    '\u00f5': 'o',   # õ
    '\u00e2': 'a',   # â
    '\u00ed': 'i',   # í
    '\u00fa': 'u',   # ú
    '\u00f3': 'o',   # ó
    '\u00e0': 'a',   # à
})


def _s(text) -> str:
    """Sanitize a value for fpdf 1.7.2 (Latin-1 only).
    Applies known substitutions then drops any remaining non-Latin-1 chars."""
    s = str(text) if text is not None else ''
    s = s.translate(_LATIN1_SUBS)
    return s.encode('latin-1', errors='ignore').decode('latin-1')


def _set_fine_black_lines(pdf: FPDF):
    """Set draw color to black and line width to hairline for all table borders."""
    pdf.set_draw_color(*LINE_COLOR)
    pdf.set_line_width(LINE_W)


def add_header(pdf: FPDF):
    """Blue header bar at top of every page."""
    pdf.set_fill_color(*HEADER_BLUE)
    pdf.rect(0, 0, 210, 18, 'F')

    pdf.set_font("Arial", 'B', 9)
    pdf.set_text_color(*HEADER_TEXT)
    pdf.set_xy(10, 4)
    pdf.cell(100, 6, "VESTAS  |  Blade Repair Materials Planner", 0, 0, 'L')

    pdf.set_font("Arial", '', 8)
    pdf.set_xy(120, 4)
    pdf.cell(80, 6, f"Confidential  |  {time.strftime('%d/%m/%Y')}", 0, 0, 'R')

    pdf.set_font("Arial", 'B', 11)
    pdf.set_text_color(*HEADER_TEXT)
    pdf.set_xy(10, 10)
    pdf.cell(190, 6, "BLADE REPAIR MATERIALS PLANNER", 0, 0, 'C')

    # Reset text color back to black after header
    pdf.set_text_color(*DARK_TEXT)


def add_footer(pdf: FPDF):
    pdf.set_y(-12)
    pdf.set_font("Arial", '', 7)
    pdf.set_text_color(*LIGHT_TEXT)
    pdf.cell(0, 6,
             f"Page {pdf.page_no()}  |  "
             f"Generated {time.strftime('%d/%m/%Y %H:%M')}  |  "
             "Vestas Internal Use",
             0, 0, 'C')


def draw_info_block(pdf: FPDF, data: BOMPayload):
    _set_fine_black_lines(pdf)
    ROW_H = 5.5

    # Title line
    pdf.set_xy(10, 24)
    pdf.set_font("Arial", 'B', 10)
    pdf.set_text_color(*DARK_TEXT)
    pdf.cell(190, 7,
             _s(f"Turbine: {data.turbine_model}  |  Zone: {data.blade_zone}"),
             0, 1, 'L')

    pdf.set_font("Courier", '', 9)
    pdf.set_text_color(*DARK_TEXT)

    def two_col(left, right, shade=False):
        pdf.set_fill_color(*(ROW_ALT if shade else ROW_WHITE))
        pdf.cell(95, ROW_H, _s(f"  {left}"),  1, 0, 'L', fill=True)
        pdf.cell(95, ROW_H, _s(f"  {right}"), 1, 1, 'L', fill=True)

    def section_label(text):
        pdf.set_fill_color(*COL_HDR_BG)
        pdf.set_font("Arial", 'B', 8)
        pdf.set_text_color(*DARK_TEXT)
        pdf.cell(190, ROW_H, f"  {text}", 1, 1, 'L', fill=True)
        pdf.set_font("Courier", '', 9)

    two_col(f"Blade SN: {data.blade_sn or '-'}",
            f"SO: {data.service_order or '-'}", shade=False)
    two_col(f"CIR Number: {data.cir_number or '-'}",
            f"Damage Category: {data.damage_cat}", shade=True)

    section_label("SPANWISE")
    span_length = abs(data.rend - data.rstart)
    pdf.set_fill_color(*ROW_WHITE)
    pdf.cell(63, ROW_H, f"  Rstart: {data.rstart:.1f} mm",  1, 0, 'L', fill=True)
    pdf.cell(63, ROW_H, f"  Length: {span_length:.1f} mm",  1, 0, 'L', fill=True)
    pdf.cell(64, ROW_H, f"  Width: {data.width:.1f} mm",    1, 1, 'L', fill=True)

    section_label("CHORDWISE")
    pdf.set_fill_color(*ROW_WHITE)
    pdf.cell(95, ROW_H, f"  Reference: {data.chord_ref or 'LE'}", 1, 0, 'L', fill=True)
    pdf.cell(95, ROW_H, f"  X: {(data.x1 or 0):.1f} mm",         1, 1, 'L', fill=True)

    two_col(f"Days of Repair: {data.days}",
            f"Region: {data.blade_zone}", shade=True)

    if data.estimated_days is not None:
        repair_kind = "External" if data.is_external else "Internal"
        two_col(f"Estimated Duration: {data.estimated_days} day(s)",
                f"Repair Type: {repair_kind}", shade=False)

    # Damage description is ALWAYS rendered, even if empty — placeholder dash
    # keeps the report layout stable for downstream readers.
    pdf.ln(1)
    pdf.set_font("Arial", 'B', 8)
    pdf.set_text_color(*MID_TEXT)
    pdf.cell(190, 5, "  DAMAGE DESCRIPTION:", 0, 1)
    pdf.set_font("Arial", '', 8)
    pdf.set_text_color(*DARK_TEXT)
    description_text = _s((data.damage_description or "").strip() or "-")
    pdf.multi_cell(190, 4.5, f"  {description_text}")

    pdf.ln(4)


_DOC_REF_FIELDS = [
    ("version",   "Version"),
    ("final",     "Blade Final"),
    ("finish",    "Blade Finish"),
    ("bonding",   "Blade Bonding"),
    ("assembled", "Blade Assembled"),
    ("shellWW",   "Shell Layup WW"),
    ("shellLW",   "Shell Layup LW"),
    ("web",       "Web"),
]


def draw_doc_refs(pdf: FPDF, data: BOMPayload):
    """Optional drawing-reference block (from BLADE_DOCUMENT_REFERENCES)."""
    refs = data.doc_refs
    if not refs:
        return
    pdf.set_font("Arial", 'B', 8.5)
    pdf.set_fill_color(*CAT_HDR_BG)
    pdf.set_text_color(*DARK_TEXT)
    pdf.cell(190, 6, "  DRAWING REFERENCES", 1, 1, 'L', fill=True)
    pdf.set_font("Arial", '', 8)
    for idx, (key, label) in enumerate(_DOC_REF_FIELDS):
        val = str(refs.get(key) or "-")
        pdf.set_fill_color(*(ROW_ALT if idx % 2 else ROW_WHITE))
        pdf.cell(60, 5, f"  {label}", 1, 0, 'L', fill=True)
        pdf.cell(130, 5, _s(val)[:78], 1, 1, 'L', fill=True)
    pdf.ln(4)


# 10 general field rules (945550 §9) — mirrors FIELD_RULES in static/data.js.
_FIELD_RULES = [
    (1,  "Surface clean before laminating",          "After peel-ply/abrasion: max 3h exposed"),
    (2,  "Fibre aligned per drawing",                "Duplicate removed-material orientation, no bumps"),
    (3,  "Minimum overlap",                          "5% of fabric weight (g/m2) in mm. Ex: 600gsm = 30mm"),
    (4,  "First ply = smallest (multi-layer)",       "In multi-layer repair, start with the smallest"),
    (5,  "Vacuum min. 0.8 bar",                      "Vacuum consolidation whenever possible"),
    (6,  "Cure with heating blanket",                "Per 0042-5383. Thermal sensor under blanket advised"),
    (7,  "Wet the surface with resin before fibre",  "Substrate wet-out mandatory"),
    (8,  "Rounded corners on laminates",             "Cut fibre with round corners"),
    (9,  "Material-surface temp diff <= 5C",         "Material and blade must be close in temperature"),
    (10, "Grinding only after full cure (95%)",      "Safety: exposure to uncured chemicals"),
]


def draw_field_rules(pdf: FPDF):
    """Optional field-rules checklist appendix (945550 §9)."""
    pdf.add_page()
    add_header(pdf)
    pdf.set_y(24)
    _set_fine_black_lines(pdf)
    pdf.set_font("Arial", 'B', 8.5)
    pdf.set_fill_color(*CAT_HDR_BG)
    pdf.set_text_color(*DARK_TEXT)
    pdf.cell(190, 6, "  FIELD RULES CHECKLIST (945550 par.9)", 1, 1, 'L', fill=True)
    pdf.set_font("Arial", 'B', 7.5)
    pdf.set_fill_color(*COL_HDR_BG)
    pdf.cell(8,  5.5, "OK", 1, 0, 'C', fill=True)
    pdf.cell(10, 5.5, "#",  1, 0, 'C', fill=True)
    pdf.cell(80, 5.5, "RULE", 1, 0, 'L', fill=True)
    pdf.cell(92, 5.5, "DETAIL", 1, 1, 'L', fill=True)
    pdf.set_font("Arial", '', 7.5)
    for idx, (n, rule, detail) in enumerate(_FIELD_RULES):
        pdf.set_fill_color(*(ROW_ALT if idx % 2 else ROW_WHITE))
        pdf.cell(8,  6, "", 1, 0, 'C', fill=True)  # empty checkbox
        pdf.cell(10, 6, str(n), 1, 0, 'C', fill=True)
        pdf.cell(80, 6, _s(rule)[:52], 1, 0, 'L', fill=True)
        pdf.cell(92, 6, _s(detail)[:60], 1, 1, 'L', fill=True)
    pdf.ln(4)


def _cat_col_headers(pdf: FPDF):
    """Draw column headers for item tables — light grey background, black text, fine borders."""
    _set_fine_black_lines(pdf)
    pdf.set_fill_color(*COL_HDR_BG)
    pdf.set_font("Arial", 'B', 7.5)
    pdf.set_text_color(*DARK_TEXT)
    pdf.cell(28,  5.5, "SAP IN",      1, 0, 'C', fill=True)
    pdf.cell(122, 5.5, "DESCRIPTION", 1, 0, 'C', fill=True)
    pdf.cell(20,  5.5, "QTY",         1, 0, 'C', fill=True)
    pdf.cell(20,  5.5, "UNIT",        1, 1, 'C', fill=True)


def draw_category(pdf: FPDF, phase: str, items: list):
    _set_fine_black_lines(pdf)

    # Check if we need a new page (enough room for header + at least 3 rows)
    if pdf.get_y() > 297 - 15 - (6 + 5.5 + 3 * 5):
        pdf.add_page()
        add_header(pdf)
        pdf.set_y(24)
        _set_fine_black_lines(pdf)

    # Category header — slightly darker grey, bold text, black borders
    pdf.set_fill_color(*CAT_HDR_BG)
    pdf.set_font("Arial", 'B', 8.5)
    pdf.set_text_color(*DARK_TEXT)
    pdf.cell(190, 6, f"  {phase.upper()}", 1, 1, 'L', fill=True)

    _cat_col_headers(pdf)

    pdf.set_font("Arial", '', 7.5)
    pdf.set_text_color(*DARK_TEXT)

    ROW_H = 5

    for idx, item in enumerate(items):
        # Page break with continuation header
        if pdf.get_y() > 275:
            pdf.add_page()
            add_header(pdf)
            pdf.set_y(24)
            _set_fine_black_lines(pdf)
            pdf.set_fill_color(*CAT_HDR_BG)
            pdf.set_font("Arial", 'B', 8.5)
            pdf.set_text_color(*DARK_TEXT)
            pdf.cell(190, 6, f"  {phase.upper()} (cont.)", 1, 1, 'L', fill=True)
            _cat_col_headers(pdf)
            pdf.set_font("Arial", '', 7.5)
            pdf.set_text_color(*DARK_TEXT)

        shade = idx % 2 == 1
        pdf.set_fill_color(*(ROW_ALT if shade else ROW_WHITE))
        pdf.cell(28,  ROW_H, _s(item.get('sap',  '-'))[:14], 1, 0, 'C', fill=True)
        pdf.cell(122, ROW_H, _s(item.get('desc', ''))[:72],  1, 0, 'L', fill=True)
        pdf.cell(20,  ROW_H, _s(item.get('qty',   0)),        1, 0, 'C', fill=True)
        pdf.cell(20,  ROW_H, _s(item.get('unit',  '')),        1, 1, 'C', fill=True)

    pdf.ln(2)


@app.post("/api/generate-pdf")
async def generate_pdf(data: BOMPayload):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=False)

    pdf.add_page()
    add_header(pdf)
    draw_info_block(pdf, data)
    draw_doc_refs(pdf, data)

    phases: dict = {}
    for item in data.items:
        phases.setdefault(item.get('phase', 'Other'), []).append(item)

    CATEGORY_ORDER = [
        'Fabrics',
        'Chemicals',
        'Consumable Tools',
        'Consumables',
        'Tools',
        'Consumable Protection Equipment',
    ]

    for phase in CATEGORY_ORDER:
        if phase in phases:
            draw_category(pdf, phase, phases[phase])
    for phase, items in phases.items():
        if phase not in CATEGORY_ORDER:
            draw_category(pdf, phase, items)

    # Optional field-rules checklist appendix
    if data.include_field_rules:
        draw_field_rules(pdf)

    # Add footer to all pages
    total_pages = pdf.page_no()
    for i in range(1, total_pages + 1):
        pdf.page = i
        add_footer(pdf)

    # Sanitize turbine and service order for filename
    def _safe(s: str):
        if not s:
            return 'UNKNOWN'
        # replace spaces with underscore and remove unsafe chars
        t = re.sub(r"\s+", '_', s.strip())
        t = re.sub(r"[^A-Za-z0-9_\-]", '', t)
        return t

    turbine_safe = _safe(data.turbine_model)
    so_safe = _safe(data.service_order or '')
    date_str = time.strftime('%Y%m%d')
    title_safe = _safe(data.report_title or '')
    filename_short = (f"{title_safe}.pdf" if title_safe and title_safe != 'UNKNOWN'
                      else f"BOM_Report_{turbine_safe}_{so_safe}_{date_str}.pdf")

    # Render the PDF into memory so the response works on stateless cloud
    # hosts (Render) where the local disk is ephemeral.
    pdf_bytes = bytes(pdf.output())

    # ── Optional disk persistence + audit trail ────────────────────────
    # On local/OneDrive deployments we keep writing a copy of the PDF and a
    # matching JSON audit record to disk. Set BOM_PERSIST=0 (done on Render)
    # to skip all disk writes. Disk failures never break the response.
    if os.environ.get('BOM_PERSIST', '1') != '0':
      try:
        # BOM_OUT_DIR env var is set by INICIAR_CALCULADORA.bat to the user's own
        # Documents folder so that shared/read-only OneDrive deployments work for
        # every user without requiring write access to the application folder.
        out_dir = os.environ.get('BOM_OUT_DIR') or os.path.join(os.getcwd(), 'BOM_Reports_Generated')
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, filename_short)
        with open(out_path, 'wb') as pf:
            pf.write(pdf_bytes)

        audit_dir = os.path.join(out_dir, 'audit')
        os.makedirs(audit_dir, exist_ok=True)
        base_no_ext = os.path.splitext(filename_short)[0]
        audit_path = os.path.join(audit_dir, f"{base_no_ext}.json")

        record = {
            "schema_version": 1,
            "pdf_filename": filename_short,
            "generated_at_utc": datetime.now(timezone.utc).isoformat(timespec='seconds').replace('+00:00', 'Z'),
            "calc_engine_version": "REV05",
            "inputs": {
                "turbine_model": data.turbine_model,
                "blade_type": data.blade_type,
                "damage_cat": data.damage_cat,
                "blade_zone": data.blade_zone,
                "blade_sn": data.blade_sn or "",
                "service_order": data.service_order or "",
                "cir_number": data.cir_number or "",
                "damage": {
                    "rstart": data.rstart,
                    "rend": data.rend,
                    "length": data.length,
                    "width": data.width,
                    "x1": data.x1,
                    "chord_ref": data.chord_ref,
                },
                "damage_description": data.damage_description or "",
                "days_of_repair": data.days,
                # repair_steps + layers come from the frontend audit_inputs blob
                "repair_steps": (data.audit_inputs or {}).get("repair_steps", {}),
                "layers": (data.audit_inputs or {}).get("layers", []),
            },
            "outputs": {
                "layup_summary": (data.audit_outputs or {}).get("layup_summary", {}),
                "items_by_phase": _group_items_by_phase(data.items),
                "totals": {
                    "total_items": len(data.items),
                    **(data.audit_outputs or {}).get("totals", {}),
                },
            },
        }

        with open(audit_path, 'w', encoding='utf-8') as f:
            json.dump(record, f, ensure_ascii=False, indent=2)
      except Exception as e:
        # Disk persistence/audit must never break the PDF response. Log and continue.
        try:
            with open(os.path.join(os.getcwd(), 'server.log'), 'a', encoding='utf-8') as lf:
                lf.write(f"[{datetime.now().isoformat()}] pdf-persist-failed: {e}\n")
        except Exception:
            pass

    return Response(
        content=pdf_bytes,
        media_type='application/pdf',
        headers={"Content-Disposition": f'attachment; filename="{filename_short}"'},
    )


@app.post("/api/import-bom-excel")
async def import_bom_excel(request: Request):
    """Lê a aba INPUTS de um Excel gerado pelo app (marcador BRMP_REEDIT_V1) e
    devolve as entradas do cálculo em JSON, para o front reabrir e reeditar.
    Recebe os bytes crus do arquivo (sem multipart — sem dependência extra)."""
    raw = await request.body()
    if not raw:
        return Response(content='{"error":"empty file"}', media_type="application/json", status_code=400)
    try:
        wb = load_workbook(BytesIO(raw), data_only=True)
    except Exception:
        return Response(content='{"error":"file is not a valid .xlsx"}', media_type="application/json", status_code=400)
    if "INPUTS" not in wb.sheetnames:
        return Response(content='{"error":"no INPUTS sheet — this file was not generated by an app version with re-editing"}',
                        media_type="application/json", status_code=422)
    ws = wb["INPUTS"]
    if str(ws.cell(1, 1).value or "").strip() != "BRMP_REEDIT_V1":
        return Response(content='{"error":"BRMP_REEDIT_V1 marker missing from the INPUTS sheet"}',
                        media_type="application/json", status_code=422)

    meta, layers, steps = {}, [], {}
    section = "meta"
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None or str(a).strip() == "":
            continue
        a = str(a).strip()
        if a == "[LAYERS]":
            section = "layers_header"; continue
        if a == "[STEPS]":
            section = "steps_header"; continue
        if section == "meta":
            meta[a] = ws.cell(r, 2).value
        elif section == "layers_header":
            section = "layers"  # pula a linha de cabeçalho (#, Material, ...)
        elif section == "layers":
            def _num(c):
                v = ws.cell(r, c).value
                return None if v is None or str(v).strip() == "" else float(v)
            layers.append({
                "materialType": str(ws.cell(r, 2).value or "").strip(),
                "gsm": str(ws.cell(r, 3).value or "").strip(),
                "ovR1": _num(4), "ovR2": _num(5), "ovX1": _num(6), "ovX2": _num(7),
            })
        elif section == "steps_header":
            section = "steps"  # pula a linha Step | Qty
        elif section == "steps":
            try:
                steps[a] = int(ws.cell(r, 2).value or 0)
            except (TypeError, ValueError):
                steps[a] = 0

    def _s(key, default=""):
        v = meta.get(key)
        return default if v is None or str(v).strip() in ("", "-") else str(v).strip()
    def _f(key, default=0):
        try:
            return float(meta.get(key))
        except (TypeError, ValueError):
            return default

    out = {
        "blade": _s("Blade"), "region": _s("Region", "Middle"),
        "length": _f("Length (mm)"), "width": _f("Width (mm)"),
        "days": int(_f("Days", 5) or 5),
        "isExternal": _s("External", "no").lower() == "yes",
        "paint": {"base": _s("Paint base", "RAL7035"), "stripe": _s("Paint stripe") or None},
        "so": _s("SO"), "cir": _s("CIR"), "title": _s("Title"),
        "layers": [l for l in layers if l["materialType"]],
        "steps": steps,
    }
    return Response(content=json.dumps(out), media_type="application/json")


@app.post("/api/generate-excel")
async def generate_excel(data: BOMPayload):
    """Generate an .xlsx bill of materials mirroring the PDF layout.
    Returned as an in-memory download so it works on stateless cloud hosts
    (Render) where the local disk is ephemeral."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    # ── Styles ──────────────────────────────────────────────────
    blue_fill = PatternFill("solid", fgColor="143A5F")
    cat_fill = PatternFill("solid", fgColor="E1E1E1")
    hdr_fill = PatternFill("solid", fgColor="EBEBEB")
    white_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    bold = Font(name="Calibri", size=10, bold=True)
    normal = Font(name="Calibri", size=10)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Column widths: SAP | DESCRIPTION | QTY | UNIT
    widths = [16, 70, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title banner ────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "VESTAS BLADES  |  Blade Repair Materials Planner"
    c.fill = blue_fill
    c.font = white_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── Info block ──────────────────────────────────────────────
    span_length = abs(data.rend - data.rstart)
    info_rows = [
        ("Turbine model", data.turbine_model, "Region", data.blade_zone),
        ("Service Order", data.service_order or "-", "CIR Number", data.cir_number or "-"),
        ("Length (mm)", f"{span_length:.0f}", "Width (mm)", f"{data.width:.0f}"),
        ("Days of repair", str(data.days), "Generated", time.strftime("%d/%m/%Y")),
    ]
    if data.estimated_days is not None:
        info_rows.append((
            "Estimated duration", f"{data.estimated_days} day(s)",
            "Repair type", "External" if data.is_external else "Internal",
        ))
    # Empilhamento e etapas na cara da lista: quem revisa vê "Vacuum 6" sem
    # abrir a aba INPUTS (revisão Reynosa ago/2026 — listas de erosão de 30×20
    # mm saíram com 12 camadas e HLU=5 carregados da lesão anterior).
    if data.audit_inputs:
        ai = data.audit_inputs
        stack_counts: dict = {}
        for ly in (ai.get("layers") or []):
            if not ly.get("materialType"):
                continue
            key = ly["materialType"] + (f" {ly['gsm']}" if ly.get("gsm") else "")
            stack_counts[key] = stack_counts.get(key, 0) + 1
        stack_txt = " · ".join(f"{n}× {k}" if n > 1 else k for k, n in stack_counts.items()) or "-"
        st = ai.get("repair_steps") or {}
        hlu, inf = int(st.get("HLU") or 0), int(st.get("Infusion") or 0)
        step_bits = [f"Vacuum {hlu + inf} (HLU {hlu} + Infusion {inf})"]
        for k in ("Weighing", "Painting", "Bonding", "LEP", "Grinding", "Cleaning"):
            if int(st.get(k) or 0):
                step_bits.append(f"{k} {int(st[k])}")
        info_rows.append(("Layup stack", stack_txt, "Repair steps", " · ".join(step_bits)))
    r = 3
    for k1, v1, k2, v2 in info_rows:
        ws.cell(r, 1, k1).font = bold
        ws.cell(r, 2, str(v1)).font = normal
        ws.cell(r, 3, k2).font = bold
        ws.cell(r, 4, str(v2)).font = normal
        r += 1
    r += 1

    # ── Items grouped by phase ──────────────────────────────────
    phases: dict = {}
    for item in data.items:
        phases.setdefault(item.get("phase", "Other"), []).append(item)

    CATEGORY_ORDER = [
        "Fabrics", "Chemicals", "Consumable Tools",
        "Consumables", "Tools", "Consumable Protection Equipment",
    ]
    ordered = [p for p in CATEGORY_ORDER if p in phases]
    ordered += [p for p in phases if p not in CATEGORY_ORDER]

    def _write_phase(sheet, row, phase, items):
        """Escreve um bloco de categoria (cabeçalho + colunas + linhas) e
        devolve a próxima linha livre, já com a linha de espaço."""
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = sheet.cell(row, 1, phase.upper())
        cell.fill = cat_fill
        cell.font = bold
        cell.alignment = left
        for col in range(1, 5):
            sheet.cell(row, col).border = border
        row += 1

        for col, label in enumerate(["SAP IN", "DESCRIPTION", "QTY", "UNIT"], start=1):
            cell = sheet.cell(row, col, label)
            cell.fill = hdr_fill
            cell.font = bold
            cell.alignment = center
            cell.border = border
        row += 1

        for item in items:
            sheet.cell(row, 1, str(item.get("sap", "-"))).alignment = center
            sheet.cell(row, 2, str(item.get("desc", ""))).alignment = left
            qcell = sheet.cell(row, 3, item.get("qty", 0))
            qcell.alignment = center
            sheet.cell(row, 4, str(item.get("unit", ""))).alignment = center
            for col in range(1, 5):
                cell = sheet.cell(row, col)
                cell.font = normal
                cell.border = border
            row += 1
        return row + 1  # blank spacer row

    # Ferramentas (equipamento durável — esmerilhadeira, aspirador, balança…)
    # vão para a aba TOOLS, separada do material de compra. "Consumable Tools"
    # (lixas, pratos, discos) é consumível comprado por reparo e fica no BOM.
    # Pedido do time em set/2026: a lista de compra não deve misturar
    # ferramenta com material.
    TOOLS_SHEET_PHASES = ("Tools",)
    for phase in ordered:
        if phase in TOOLS_SHEET_PHASES:
            continue
        r = _write_phase(ws, r, phase, phases[phase])

    tools_phases = [p for p in ordered if p in TOOLS_SHEET_PHASES]
    if tools_phases:
        wt = wb.create_sheet("TOOLS")
        for i, w in enumerate(widths, start=1):
            wt.column_dimensions[get_column_letter(i)].width = w
        wt.merge_cells("A1:D1")
        ct = wt["A1"]
        ct.value = "VESTAS BLADES  |  Blade Repair Materials Planner  —  TOOLS"
        ct.fill = blue_fill
        ct.font = white_font
        ct.alignment = Alignment(horizontal="center", vertical="center")
        wt.row_dimensions[1].height = 24
        rt = 3
        for k1, v1, k2, v2 in info_rows[:2]:  # Turbine/Region, SO/CIR
            wt.cell(rt, 1, k1).font = bold
            wt.cell(rt, 2, str(v1)).font = normal
            wt.cell(rt, 3, k2).font = bold
            wt.cell(rt, 4, str(v2)).font = normal
            rt += 1
        rt += 1
        for phase in tools_phases:
            rt = _write_phase(wt, rt, phase, phases[phase])

    # ── Drawing references (optional) ───────────────────────────
    if data.doc_refs:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        cell = ws.cell(r, 1, "DRAWING REFERENCES")
        cell.fill = cat_fill
        cell.font = bold
        cell.alignment = left
        for col in range(1, 5):
            ws.cell(r, col).border = border
        r += 1
        for key, label in _DOC_REF_FIELDS:
            ws.cell(r, 1, label).font = bold
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            ws.cell(r, 2, str(data.doc_refs.get(key) or "-")).font = normal
            r += 1
        r += 1

    # ── Aba INPUTS: entradas completas do cálculo, para reedição ─────────────
    # Formato estável (marcador BRMP_REEDIT_V1) que /api/import-bom-excel lê de
    # volta — o Excel entregue vira também o "arquivo de projeto" do reparo.
    if data.audit_inputs:
        ai = data.audit_inputs
        ws2 = wb.create_sheet("INPUTS")
        # Aba oculta de propósito: quem recebe o Excel vê só o BOM; quem conhece
        # o sistema faz Unhide (botão direito na aba) ou usa "Reabrir de Excel",
        # que lê a aba oculta normalmente.
        ws2.sheet_state = "hidden"
        ws2.cell(1, 1, "BRMP_REEDIT_V1").font = bold
        meta_rows = [
            ("Blade", data.turbine_model), ("Region", data.blade_zone),
            ("Length (mm)", data.length), ("Width (mm)", data.width),
            ("Days", data.days), ("External", "yes" if data.is_external else "no"),
            ("Paint base", (ai.get("paint") or {}).get("base") or "-"),
            ("Paint stripe", (ai.get("paint") or {}).get("stripe") or "-"),
            ("SO", data.service_order or "-"), ("CIR", data.cir_number or "-"),
            ("Title", data.report_title or "-"),
        ]
        r2 = 2
        for k, v in meta_rows:
            ws2.cell(r2, 1, k).font = bold
            ws2.cell(r2, 2, v).font = normal
            r2 += 1
        r2 += 1
        ws2.cell(r2, 1, "[LAYERS]").font = bold
        r2 += 1
        for j, h in enumerate(["#", "Material", "GSM", "ovR1", "ovR2", "ovX1", "ovX2"], 1):
            ws2.cell(r2, j, h).font = bold
        r2 += 1
        for ly in (ai.get("layers") or []):
            ws2.cell(r2, 1, ly.get("order"))
            ws2.cell(r2, 2, ly.get("materialType"))
            ws2.cell(r2, 3, ly.get("gsm") or "")
            for j, k in enumerate(["ovR1", "ovR2", "ovX1", "ovX2"], 4):
                v = ly.get(k)
                if v is not None and v != "":
                    ws2.cell(r2, j, v)
            r2 += 1
        r2 += 1
        ws2.cell(r2, 1, "[STEPS]").font = bold
        r2 += 1
        ws2.cell(r2, 1, "Step").font = bold
        ws2.cell(r2, 2, "Qty").font = bold
        r2 += 1
        for k, v in (ai.get("repair_steps") or {}).items():
            ws2.cell(r2, 1, k)
            ws2.cell(r2, 2, v)
            r2 += 1
        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 22

    # ── Write to in-memory buffer ───────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    def _safe(s: str):
        if not s:
            return "UNKNOWN"
        t = re.sub(r"\s+", "_", s.strip())
        return re.sub(r"[^A-Za-z0-9_\-]", "", t)

    title_safe = _safe(data.report_title or "")
    filename = (f"{title_safe}.xlsx" if title_safe and title_safe != "UNKNOWN"
                else f"BOM_Report_{_safe(data.turbine_model)}_{_safe(data.service_order or '')}_{time.strftime('%Y%m%d')}.xlsx")

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class ScarfPayload(BaseModel):
    meta: Dict[str, Any] = {}
    columns: list = []
    rows: list = []


@app.post("/api/scarf-export")
async def scarf_export(data: ScarfPayload, fmt: str = "xlsx"):
    """Export the layer staggering (scarfing / escalonamento) coordinate
    table as .xlsx or .pdf. In-memory / stateless."""
    meta = data.meta or {}
    cols = data.columns or []
    rows = data.rows or []

    def _safe(s):
        t = re.sub(r"\s+", "_", str(s or "").strip())
        return re.sub(r"[^A-Za-z0-9_\-]", "", t) or "UNKNOWN"

    fname_base = f"Escalonamento_{_safe(meta.get('blade'))}_{_safe(meta.get('service_order'))}"

    info_pairs = [
        ("Blade", meta.get("blade", "-"), "Region", meta.get("region", "-")),
        ("Damage L x W", f"{meta.get('length','-')} x {meta.get('width','-')} mm",
         "Z start", f"{meta.get('z0', 0)} mm"),
        ("Service Order", meta.get("service_order", "-") or "-",
         "Chord ref", meta.get("chord_ref", "-")),
    ]

    if fmt == "pdf":
        pdf = FPDF(orientation="L")
        pdf.set_auto_page_break(auto=True, margin=12)
        pdf.add_page()
        # header bar
        pdf.set_fill_color(*HEADER_BLUE)
        pdf.rect(0, 0, 297, 16, "F")
        pdf.set_font("Arial", "B", 11)
        pdf.set_text_color(*HEADER_TEXT)
        pdf.set_xy(10, 5)
        pdf.cell(0, 6, _s("VESTAS BLADES  |  Layer Staggering (Escalonamento)"), 0, 1, "L")
        pdf.set_text_color(*DARK_TEXT)
        pdf.ln(6)
        # info block
        pdf.set_font("Arial", "", 9)
        for a, b, c, d in info_pairs:
            pdf.cell(35, 6, _s(f"{a}:"), 0, 0)
            pdf.set_font("Arial", "B", 9); pdf.cell(75, 6, _s(b), 0, 0); pdf.set_font("Arial", "", 9)
            pdf.cell(30, 6, _s(f"{c}:"), 0, 0)
            pdf.set_font("Arial", "B", 9); pdf.cell(60, 6, _s(d), 0, 1); pdf.set_font("Arial", "", 9)
        pdf.ln(3)
        # table
        _set_fine_black_lines(pdf)
        n = len(cols) or 1
        total_w = 277
        cw = total_w / n
        pdf.set_fill_color(*COL_HDR_BG)
        pdf.set_font("Arial", "B", 8)
        for c in cols:
            pdf.cell(cw, 7, _s(c), 1, 0, "C", fill=True)
        pdf.ln()
        pdf.set_font("Arial", "", 8)
        for i, row in enumerate(rows):
            pdf.set_fill_color(*(ROW_ALT if i % 2 else ROW_WHITE))
            for v in row:
                if isinstance(v, float) and v == int(v):
                    v = int(v)
                pdf.cell(cw, 6, _s(v), 1, 0, "C", fill=True)
            pdf.ln()
        out = bytes(pdf.output())
        return Response(
            content=out, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname_base}.pdf"'},
        )

    # default: xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "Escalonamento"
    blue_fill = PatternFill("solid", fgColor="143A5F")
    hdr_fill = PatternFill("solid", fgColor="EBEBEB")
    white_font = Font(bold=True, color="FFFFFF", size=12)
    bold = Font(bold=True, size=10)
    normal = Font(size=10)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    ncols = max(len(cols), 4)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, "VESTAS BLADES  |  Layer Staggering (Escalonamento)")
    c.fill = blue_fill; c.font = white_font; c.alignment = center
    ws.row_dimensions[1].height = 22

    r = 3
    for a, b, cc, d in info_pairs:
        ws.cell(r, 1, a).font = bold
        ws.cell(r, 2, str(b)).font = normal
        if ncols >= 4:
            ws.cell(r, 3, cc).font = bold
            ws.cell(r, 4, str(d)).font = normal
        r += 1
    r += 1

    for j, col in enumerate(cols, start=1):
        cell = ws.cell(r, j, col)
        cell.fill = hdr_fill; cell.font = bold; cell.alignment = center; cell.border = border
    r += 1
    for row in rows:
        for j, v in enumerate(row, start=1):
            if isinstance(v, float) and v == int(v):
                v = int(v)
            cell = ws.cell(r, j, v)
            cell.font = normal; cell.alignment = center; cell.border = border
        r += 1

    for j in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname_base}.xlsx"'},
    )


def _group_items_by_phase(items: list) -> Dict[str, list]:
    """Group BOM items by their 'phase' field for the audit record."""
    grouped: Dict[str, list] = {}
    for it in items:
        phase = it.get('phase', 'Other') if isinstance(it, dict) else 'Other'
        grouped.setdefault(phase, []).append(it)
    return grouped